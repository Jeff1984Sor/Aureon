from django.shortcuts import get_object_or_404, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q
from django.core.mail import send_mail
from django.template.loader import render_to_string
from collections import OrderedDict
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from weasyprint import HTML
from django.contrib.auth import get_user_model

# Modelos e Formulários do App
from .models import (
    Caso, Status, Cobertura, Motivo, Timesheet, Cliente, Advogado, Analista,
    FaseWorkflow, RegraWorkflow, Tarefa, TipoTarefa, Produto, Andamento
)
from .forms import CasoCreateForm, CasoUpdateForm, AndamentoForm, TarefaForm, TimesheetForm, TarefaConclusaoForm

Usuario = get_user_model()

# ==============================================================================
# VIEWS DE CASOS (CRUD)
# ==============================================================================

class CasoListView(LoginRequiredMixin, ListView):
    model = Caso
    template_name = 'casos/caso_list.html'
    context_object_name = 'casos'
    paginate_by = 10


class CasoCreateView(LoginRequiredMixin, CreateView):
    model = Caso
    form_class = CasoCreateForm
    template_name = 'casos/caso_form.html'
    success_url = reverse_lazy('caso_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Adicionar Novo Caso'
        form = context.get('form')
        if form:
            
            form.fields['cliente'].queryset = Cliente.objects.filter(empresa=empresa_usuario)
            form.fields['advogado_responsavel'].queryset = Advogado.objects.filter(empresa=empresa_usuario)
            form.fields['status'].queryset = Status.objects.filter(empresa=empresa_usuario)
            form.fields['produto'].queryset = Produto.objects.filter(empresa=empresa_usuario)
            form.fields['cobertura'].queryset = Cobertura.objects.filter(empresa=empresa_usuario)
            form.fields['motivo'].queryset = Motivo.objects.filter(empresa=empresa_usuario)
            form.fields['analista'].queryset = Analista.objects.filter(empresa=empresa_usuario)
        return context

    def form_valid(self, form):
       
        response = super().form_valid(form)
        novo_caso = self.object
        
        # Lógica de e-mail (agora que o caso foi salvo)
        # Implementar ou remover se não for usar agora
        
        return response

class CasoUpdateView(LoginRequiredMixin, UpdateView):
    model = Caso
    form_class = CasoUpdateForm
    template_name = 'casos/caso_form.html'
    success_url = reverse_lazy('caso_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Caso'
        return context

class CasoDetailView(LoginRequiredMixin, DetailView):
    model = Caso
    template_name = 'casos/caso_detail.html'
    context_object_name = 'caso'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['andamento_form'] = AndamentoForm()
        context['tarefa_form'] = TarefaForm()
        context['timesheet_form'] = TimesheetForm()
        context['conclusao_form'] = TarefaConclusaoForm()
        return context

# ==============================================================================
# VIEWS DE AÇÕES (ANDAMENTO, TAREFA, TIMESHEET)
# ==============================================================================

@login_required
def add_andamento(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk, empresa=request.user.empresa)
    if request.method == 'POST':
        form = AndamentoForm(request.POST)
        if form.is_valid():
            andamento = form.save(commit=False)
            andamento.caso = caso
            andamento.usuario_criacao = request.user
            andamento.save()
    return redirect(reverse('caso_detail', kwargs={'pk': caso_pk}) + '#andamentos-tab-pane')

@login_required
def add_tarefa(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk, empresa=request.user.empresa)
    if request.method == 'POST':
        form = TarefaForm(request.POST)
        if form.is_valid():
            tarefa = form.save(commit=False)
            tarefa.caso = caso
            if not tarefa.responsavel:
                tarefa.responsavel = request.user
            tarefa.save()
    return redirect(reverse('caso_detail', kwargs={'pk': caso_pk}) + '#tarefas-tab-pane')

@login_required
def concluir_tarefa(request, tarefa_pk):
    tarefa = get_object_or_404(Tarefa, pk=tarefa_pk, caso__empresa=request.user.empresa)
    if request.method == 'POST':
        form = TarefaConclusaoForm(request.POST, instance=tarefa)
        if form.is_valid():
            tarefa_concluida = form.save(commit=False)
            tarefa_concluida.status = 'C'
            tarefa_concluida.data_conclusao = timezone.now()
            tarefa_concluida.save()
            Andamento.objects.create(
                caso=tarefa.caso, data_andamento=timezone.now().date(),
                usuario_criacao=request.user,
                descricao=f"Tarefa Concluída: {tarefa.tipo_tarefa.nome}\nDescrição: {tarefa_concluida.descricao_conclusao}"
            )
    next_url = request.POST.get('next', reverse('caso_detail', kwargs={'pk': tarefa.caso.pk}))
    return redirect(next_url)

@login_required
def reabrir_tarefa(request, tarefa_pk):
    tarefa = get_object_or_404(Tarefa, pk=tarefa_pk, caso__empresa=request.user.empresa)
    if request.method == 'POST':
        tarefa.status = 'P'
        tarefa.data_conclusao = None
        tarefa.save()
    next_url = request.POST.get('next', reverse('caso_detail', kwargs={'pk': tarefa.caso.pk}))
    return redirect(next_url)

@login_required
def deletar_tarefa(request, tarefa_pk):
    tarefa = get_object_or_404(Tarefa, pk=tarefa_pk, caso__empresa=request.user.empresa)
    caso_pk = tarefa.caso.pk
    if request.method == 'POST':
        tarefa.delete()
    next_url = request.POST.get('next', reverse('caso_detail', kwargs={'pk': caso_pk}))
    return redirect(next_url)

@login_required
def add_timesheet(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk, empresa=request.user.empresa)
    if request.method == 'POST':
        form = TimesheetForm(request.POST)
        if form.is_valid():
            timesheet = form.save(commit=False)
            timesheet.caso = caso
            tempo_str = form.cleaned_data['tempo_str']
            horas, minutos = map(int, tempo_str.split(':'))
            timesheet.tempo = timedelta(hours=horas, minutes=minutos)
            timesheet.save()
    return redirect(reverse('caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

@login_required
def delete_timesheet(request, ts_pk):
    timesheet = get_object_or_404(Timesheet, pk=ts_pk, caso__empresa=request.user.empresa)
    caso_pk = timesheet.caso.pk
    if request.method == 'POST':
        timesheet.delete()
    return redirect(reverse('caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

class TarefaListView(LoginRequiredMixin, ListView):
    model = Tarefa
    template_name = 'casos/tarefa_list.html'
    context_object_name = 'tarefas'
    paginate_by = 20

    def get_queryset(self):
        """
        Sobrescreve o queryset padrão para filtrar tarefas pela empresa do usuário
        e aplicar os filtros de busca da URL.
        """
            
        # Começa com apenas as tarefas da empresa do usuário, otimizando com select_related
       
        # Filtro por Responsável (usuário)
        responsavel_id = self.request.GET.get('responsavel')
        if responsavel_id:
            queryset = queryset.filter(responsavel_id=responsavel_id)

        # Filtro por Data de Criação (intervalo)
        data_de = self.request.GET.get('data_de')
        if data_de:
            queryset = queryset.filter(data_criacao__date__gte=data_de)
        
        data_ate = self.request.GET.get('data_ate')
        if data_ate:
            queryset = queryset.filter(data_criacao__date__lte=data_ate)

        # Filtro por Status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        return queryset.order_by('-data_criacao')

    def get_context_data(self, **kwargs):
        """
        Adiciona dados extras ao contexto do template, como as listas
        para popular os dropdowns de filtro.
        """
        context = super().get_context_data(**kwargs)
        empresa_usuario = self.request.user.empresa
        
        # Passa a lista de usuários da mesma empresa para o filtro
        context['responsaveis'] = Usuario.objects.filter(empresa=empresa_usuario).order_by('username')
        
        # Passa as opções de status
        context['status_choices'] = Tarefa.STATUS_TAREFA_CHOICES
        
        # Passa o formulário de conclusão para o modal
        context['conclusao_form'] = TarefaConclusaoForm()
        
        return context
class TimesheetListView(LoginRequiredMixin, ListView):
    model = Timesheet
    template_name = 'casos/timesheet_list.html' # Novo template
    context_object_name = 'timesheets'
    paginate_by = 20

    
        


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_usuario = self.request.user.empresa
        context['titulo'] = "Visão Geral de Timesheet"
        context['profissionais'] = Usuario.objects.filter(empresa=empresa_usuario)
        return context
# ==============================================================================
# VIEWS DE RELATÓRIOS E PESQUISA
# ==============================================================================

def get_casos_filtrados(request):
    queryset = Caso.objects.filter(empresa=request.user.empresa).select_related('cliente', 'status', 'advogado_responsavel__user', 'analista')
    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(titulo_caso__icontains=q) | Q(numero_sinistro__icontains=q) |
            Q(cliente__nome_razao_social__icontains=q)
        )
    filtros = {
        'cliente_id': request.GET.get('cliente'),
        'status_id': request.GET.get('status'),
        'advogado_responsavel__user_id': request.GET.get('advogado'),
        'analista_id': request.GET.get('analista'),
    }
    filtros_validos = {k: v for k, v in filtros.items() if v}
    if filtros_validos:
        queryset = queryset.filter(**filtros_validos)
    return queryset.order_by('-data_entrada')

class CasoPesquisaView(LoginRequiredMixin, ListView):
    model = Caso
    template_name = 'casos/caso_pesquisa.html'
    context_object_name = 'casos'
    paginate_by = 25
    def get_queryset(self):
        return get_casos_filtrados(self.request)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_usuario = self.request.user.empresa
        context['titulo'] = "Pesquisa Avançada de Casos"
        context['clientes'] = Cliente.objects.filter(empresa=empresa_usuario)
        context['advogados'] = Advogado.objects.filter(empresa=empresa_usuario)
        context['status_list'] = Status.objects.filter(empresa=empresa_usuario)
        context['analistas'] = Analista.objects.filter(empresa=empresa_usuario)
        return context

@login_required
def exportar_casos_excel(request):
    queryset = get_casos_filtrados(request)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Relatorio de Casos'
    headers = [
        'ID do Caso', 'Título', 'Cliente', 'Status', 'Data de Entrada', 'Advogado',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for caso in queryset:
        sheet.append([
            caso.id, caso.titulo_caso, caso.cliente.nome_razao_social,
            caso.status.nome if caso.status else '', caso.data_entrada,
            str(caso.advogado_responsavel) if caso.advogado_responsavel else ''
        ])
    filename = 'Relatorio_Casos_Aureon.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

@login_required
def exportar_timesheet_excel(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk, empresa=request.user.empresa)
    timesheets = caso.timesheets.all().order_by('data_execucao')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Timesheet Caso {caso.id}'
    headers = ['Data Execução', 'Profissional', 'Tempo Gasto (HH:MM)', 'Descrição']
    sheet.append(headers)
    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font
    total_duration = timedelta()
    for ts in timesheets:
        total_duration += ts.tempo
        total_seconds = int(ts.tempo.total_seconds())
        horas = total_seconds // 3600
        minutos = (total_seconds % 3600) // 60
        tempo_formatado = f"{horas:02d}:{minutos:02d}"
        sheet.append([ts.data_execucao, ts.profissional.username, tempo_formatado, ts.descricao])
    total_total_seconds = int(total_duration.total_seconds())
    total_horas = total_total_seconds // 3600
    total_minutos = (total_total_seconds % 3600) // 60
    total_formatado = f"{total_horas:02d}:{total_minutos:02d}"
    sheet.append([])
    sheet.append(['', '', 'Total de Horas:', total_formatado])
    sheet.cell(row=sheet.max_row, column=3).font = Font(bold=True)
    sheet.cell(row=sheet.max_row, column=4).font = Font(bold=True)
    filename = f'Relatorio_Timesheet_Caso_{caso.id}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

@login_required
def gerar_caso_pdf(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk, empresa=request.user.empresa)
    html_string = render_to_string('casos/relatorio_caso_pdf.html', {'caso': caso})
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f'Relatorio_Caso_{caso.id}.pdf'
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

class KanbanView(LoginRequiredMixin, TemplateView):
    template_name = 'casos/kanban_board.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_usuario = self.request.user.empresa
        context['titulo'] = "Kanban de Casos por Fase"
        
        casos_filtrados = Caso.objects.filter(empresa=empresa_usuario)
        profissional_id = self.request.GET.get('profissional')
        if profissional_id:
            casos_filtrados = casos_filtrados.filter(advogado_responsavel__user_id=profissional_id)
        status_id = self.request.GET.get('status')
        if status_id:
            casos_filtrados = casos_filtrados.filter(status_id=status_id)
        
        fases_usadas_ids = {c.fase_atual_workflow_id for c in casos_filtrados if c.fase_atual_workflow_id}
        fases_para_exibir = FaseWorkflow.objects.filter(id__in=fases_usadas_ids, workflow__empresa=empresa_usuario).order_by('ordem')
        
        if not (profissional_id or status_id):
            fases_para_exibir = FaseWorkflow.objects.filter(workflow__empresa=empresa_usuario).order_by('workflow__nome', 'ordem')

        kanban_columns = OrderedDict()
        for fase in fases_para_exibir:
            casos_desta_fase = [c for c in casos_filtrados if c.fase_atual_workflow == fase]
            kanban_columns[fase] = casos_desta_fase
            
        context['kanban_columns'] = kanban_columns
        context['profissionais'] = Usuario.objects.filter(empresa=empresa_usuario).order_by('username')
        context['status_list'] = Status.objects.filter(empresa=empresa_usuario)
        return context

# ==============================================================================
# VIEWS AJAX (PARA OS BOTÕES '+')
# ==============================================================================

@login_required
def add_status_ajax(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            status, created = Status.objects.get_or_create(empresa=request.user.empresa, nome=nome)
            return JsonResponse({'id': status.id, 'nome': status.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)

@login_required
def add_cobertura_ajax(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            cobertura, created = Cobertura.objects.get_or_create(empresa=request.user.empresa, nome=nome)
            return JsonResponse({'id': cobertura.id, 'nome': cobertura.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)

@login_required
def add_motivo_ajax(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            motivo, created = Motivo.objects.get_or_create(empresa=request.user.empresa, nome=nome)
            return JsonResponse({'id': motivo.id, 'nome': motivo.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)

@login_required
def add_analista_ajax(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            analista, created = Analista.objects.get_or_create(empresa=request.user.empresa, nome=nome)
            return JsonResponse({'id': analista.id, 'nome': analista.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)